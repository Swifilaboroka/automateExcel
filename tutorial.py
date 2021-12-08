from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')
ws = wb.active
print(ws['A2'].value)
ws['A2'].value = "Test"
print(wb['Sayfa1'])
wb.create_sheet("YeniSayfa")
# wb.save('Grades.xlsx')
print(wb.sheetnames)

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

ws.merge_cells("A10:D20")
ws.unmerge_cells("A10:D20")

ws.insert_rows(7)  # boş satır ekleme
ws.delete_rows(7)

ws.insert_cols(2)
ws.delete_cols(2)

ws.move_range("C1:D11", rows=2, cols=2)
